"""导入引擎主流程（PRD §4~§7 串联，字段对齐数据字典）。

两段式：
1. 上传 → parse → preview（解析+校验，不写主数据）
2. 确认导入 → commit（事务写入 package/package_item/import_error/review_queue）

幂等：唯一键 = 包裹号 + SKU，默认跳过，覆盖需管理员授权。
"""

import io
import uuid
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    ImportBatch,
    ImportError,
    Package,
    PackageItem,
    ReviewQueue,
    Sku,
    VirtualRule,
)
from app.services.importer.bom import normalize_remark
from app.services.importer.parser import (
    MergedItem,
    parse_paid_at,
    parse_rows,
    split_sku,
)
from app.services.importer.validator import (
    SLA_DAYS,
    apply_virtual_rules,
    check_quantity_trap,
)


def read_excel(file_bytes: bytes) -> list[dict]:
    """读取订单模板 Excel → 行 dict 列表（跳过非字段名首行）。"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = None
    data_rows = []
    for row in rows_iter:
        if header is None:
            if row and "包裹号" in [str(c).strip() if c else "" for c in row]:
                header = [str(c).strip() if c else "" for c in row]
                continue
            else:
                continue
        data_rows.append(row)

    if header is None:
        raise ValueError("未找到表头（缺少「包裹号」列）")

    result = []
    for row in data_rows:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for i, col in enumerate(header):
            if col == "":
                continue
            val = row[i] if i < len(row) else None
            item[col] = val
        result.append(item)
    return result


def _extract_qty(rows: list[dict], merged: list[MergedItem]):
    """把每行的「商品总数」汇总到合并项的 qty。"""
    from collections import defaultdict

    qty_map = defaultdict(int)
    for row in rows:
        package_no = (row.get("包裹号") or "").strip()
        sku_cell = row.get("商品SKU")
        total = row.get("商品总数")
        try:
            total = int(total)
        except (ValueError, TypeError):
            total = 0
        for sku in split_sku(sku_cell):
            qty_map[(package_no, sku)] += total

    for m in merged:
        m.qty = qty_map.get((m.package_no, m.sku), 0)


def preview(file_bytes: bytes, db: Session) -> dict:
    """阶段一：解析 + 校验，返回预览结果（不写主数据）。"""
    rows = read_excel(file_bytes)
    if not rows:
        return {"total_rows": 0, "items": [], "errors": [], "reviews": []}

    merged, _ = parse_rows(rows)
    _extract_qty(rows, merged)

    rules = list(db.scalars(select(VirtualRule).where(VirtualRule.enabled == 1).order_by(VirtualRule.priority)))
    existing_skus = {s.sku_code for s in db.scalars(select(Sku)).all()}

    from collections import defaultdict
    items_by_package = defaultdict(list)
    for m in merged:
        items_by_package[m.package_no].append(m)

    results = {}
    preview_items = []
    error_items = []
    review_items = []

    for m in merged:
        from app.services.importer.validator import ValidationResult, _check_business, _check_structure
        vr = ValidationResult(m.package_no, m.sku)
        _check_structure(m, vr)
        _check_business(m, vr, existing_skus, {})
        apply_virtual_rules(m, vr, rules)
        results[(m.package_no, m.sku)] = vr

    check_quantity_trap(items_by_package, results)

    for m in merged:
        vr = results[(m.package_no, m.sku)]
        entry = {
            "package_no": m.package_no,
            "sku": m.sku,
            "tracking_no": m.tracking_no,
            "logistics_channel": m.logistics_method,
            "pay_time": m.paid_at,
            "remark": normalize_remark(m.remark),
            "qty": m.qty,
            "rule_action": vr.rule_action,
            "errors": vr.errors,
            "reviews": vr.reviews,
            "warnings": vr.warnings,
        }
        if vr.is_blocked:
            error_items.append(entry)
        elif vr.reviews:
            review_items.append(entry)
        else:
            preview_items.append(entry)

    return {
        "total_rows": len(rows),
        "normal_count": len(preview_items),
        "error_count": len(error_items),
        "review_count": len(review_items),
        "items": preview_items,
        "errors": error_items,
        "reviews": review_items,
    }


def commit_import(file_bytes: bytes, db: Session, operator: str = "", overwrite: bool = False) -> dict:
    """阶段二：确认导入（事务写入）。"""
    preview_result = preview(file_bytes, db)

    batch_no = f"IMP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    batch = ImportBatch(
        batch_no=batch_no,
        import_type="order",
        file_name="order.xlsx",
        total_rows=preview_result["total_rows"],
        success_rows=preview_result["normal_count"],
        error_rows=preview_result["error_count"],
        review_rows=preview_result["review_count"],
        status="done",
    )
    db.add(batch)
    db.flush()

    # SKU 映射：sku_code -> Sku（用于写入 package_item.sku_id）
    sku_map = {s.sku_code: s for s in db.scalars(select(Sku)).all()}

    # 幂等唯一键集合：package_no -> 已存在的 SKU 集合
    existing_pkg_items = {}
    for p, i in db.execute(
        select(Package, PackageItem).where(PackageItem.package_id == Package.id)
    ).all():
        existing_pkg_items.setdefault(p.package_no, set()).add(i.sku_id)

    # 写入正常明细
    for item in preview_result["items"]:
        sku = sku_map.get(item["sku"])
        if sku is None:
            continue  # 未建档 SKU 已在 preview 阶段被标记 error/review，正常明细不会出现
        # 幂等判断
        if sku.id in existing_pkg_items.get(item["package_no"], set()) and not overwrite:
            continue

        pkg = db.scalar(select(Package).where(Package.package_no == item["package_no"]))
        if not pkg:
            paid_dt = parse_paid_at(item["pay_time"])
            sla_deadline = paid_dt + timedelta(days=SLA_DAYS) if paid_dt else None
            pkg = Package(
                package_no=item["package_no"],
                tracking_no=item["tracking_no"],
                logistics_channel=item["logistics_channel"],
                pay_time=paid_dt,
                sla_deadline=sla_deadline,
                status="unassigned",
                total_qty=item["qty"],
                is_resend=1 if item["package_no"].endswith("R1") else 0,
                remark=item["remark"],
            )
            db.add(pkg)
            db.flush()

        is_virtual = 1 if item["rule_action"] == "skip" else 0
        pi = PackageItem(
            package_id=pkg.id,
            sku_id=sku.id,
            qty=item["qty"],
            is_virtual=is_virtual,
            is_marker=1 if sku.sku_type == "marker" else 0,
            picked_qty=0,
        )
        db.add(pi)

    # 写入错误报告
    for err in preview_result["errors"]:
        db.add(
            ImportError(
                batch_id=batch.id,
                row_no=0,
                package_no=err["package_no"],
                sku=err["sku"],
                error_code=";".join([e.split(" ")[0] for e in err["errors"]]),
                error_msg="; ".join(err["errors"]),
                raw_data=str(err),
            )
        )

    # 写入人工复核队列
    for rev in preview_result["reviews"]:
        db.add(
            ReviewQueue(
                batch_id=batch.id,
                package_no=rev["package_no"],
                reason_code=_reason_code(rev),
                reason_detail="; ".join(rev["reviews"]),
                status="pending",
            )
        )

    db.commit()
    return {
        "batch_no": batch_no,
        "total_rows": preview_result["total_rows"],
        "success_rows": preview_result["normal_count"],
        "error_rows": preview_result["error_count"],
        "review_rows": preview_result["review_count"],
    }


def _reason_code(rev: dict) -> str:
    """根据 review 内容映射 reason_code。"""
    text = " ".join(rev["reviews"])
    if "数量陷阱" in text:
        return "R-01"
    if "CS00-Check" in rev.get("sku", ""):
        return "R-02"
    if "未建档" in text:
        return "R-03"
    if "缺货" in text:
        return "R-04"
    if rev["package_no"].endswith("R1"):
        return "R-05"
    return "R-02"
