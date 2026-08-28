"""T3.3 期初数据切换：导入文件A（SKU+库存+库位）、文件B（BOM）、修正规则配置。

字段对齐数据字典（英文枚举 + 标准审计字段）。
用法（在 backend 容器内）：
    python -m app.init_switch
幂等：重复执行会先清空期初导入的数据再重建。
仅依赖 openpyxl。
"""

import os
from openpyxl import load_workbook
from sqlalchemy import text

from app.core.database import SessionLocal
from app.models import (
    Sku, Bom, Warehouse, Shelf, Location, Stock, StockTransaction, VirtualRule,
)

DATA_DIR = "/app"
FILE_A = os.path.join(DATA_DIR, "文件A_库存台账.xlsx")
FILE_B = os.path.join(DATA_DIR, "文件B_BOM.xlsx")

# 真实标记 SKU（以实际数据为准，感叹号数量与文档略有差异）
MARKER_SKUS = {
    "CS99-No!!!!!!!!!!!!!!!!!",
    "CS00-Check!!!!!!!!!!!!!",
    "CS000-ignore",
}

STAGING_LOCATION = "STAGING-00-00"


def read_rows(path, header_row=2):
    """读取 xlsx，返回 (columns, rows)。"""
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cols = rows[header_row - 1]
    data = rows[header_row:]
    return cols, data


def parse_location_code(code):
    parts = code.split("-")
    if len(parts) == 3:
        return parts[0], parts[1], parts[2]
    return code, "", ""


def parse_bom_segment(seg):
    idx = seg.rfind("*")
    comp = seg[:idx].strip()
    qty = int(seg[idx + 1:].strip() or 1)
    return comp, qty


def clear_init_data(db):
    for t in ["stock_transaction", "stock", "allocation", "package_item", "package",
              "bom", "sku", "location", "shelf"]:
        db.execute(text(f"TRUNCATE {t} RESTART IDENTITY CASCADE"))
    db.commit()
    print("已清空期初数据相关表")


def fix_rules(db):
    """规则配置（对齐数据字典 §7.3：rule_type/match_type/match_value/action）。"""
    want = [
        ("virtual", "exact", "YUN", "skip", 1, "运费虚拟品：免拣货、免库存校验、免面单"),
        ("marker", "prefix", "CS99", "intercept", 2, "订单不可发货警告：整包裹拦截出库"),
        ("marker", "prefix", "CS00", "manual_review", 2, "电器规格匹配错误警告：强制人工复核"),
        ("marker", "prefix", "CS000", "ignore", 2, "忽略采购拆链组合发货：采购需求/组合展开时跳过"),
    ]
    # 清空旧规则
    db.execute(text("TRUNCATE virtual_rule RESTART IDENTITY CASCADE"))
    db.commit()
    for rtype, mtype, mval, action, prio, desc in want:
        db.add(VirtualRule(rule_type=rtype, match_type=mtype, match_value=mval,
                           action=action, priority=prio, enabled=1, description=desc))
    db.commit()
    print("规则配置已写入：YUN=skip / CS99=intercept / CS00=manual_review / CS000=ignore")


def import_file_a(db):
    cols, rows = read_rows(FILE_A, header_row=2)
    print(f"文件A：{len(rows)} 行，列={cols}")

    # 仓库
    db.execute(text("TRUNCATE warehouse RESTART IDENTITY CASCADE"))
    wh = Warehouse(code="CS", name="长沙CS仓")
    db.add(wh)
    db.commit()
    db.refresh(wh)
    print(f"仓库：{wh.code} (id={wh.id})")

    # 货架（从库位码第一段提取）
    shelf_map = {}
    loc_codes = set()
    for r in rows:
        loc = r[2]
        if loc and loc != "无货架位":
            loc_codes.add(loc)
    for code in loc_codes:
        sc, _, _ = parse_location_code(code)
        shelf_map.setdefault(sc, Shelf(warehouse_id=wh.id, code=sc, name=sc, status=1))
    for s in shelf_map.values():
        db.add(s)
    db.commit()
    for s in shelf_map.values():
        db.refresh(s)
    print(f"货架：{len(shelf_map)} 个 -> {sorted(shelf_map.keys())}")

    # 库位
    loc_map = {}
    for code in loc_codes:
        sc, col, layer = parse_location_code(code)
        loc = Location(shelf_id=shelf_map[sc].id, code=code,
                       shelf_no=sc, column_no=col, layer_no=layer, status="empty")
        db.add(loc)
        loc_map[code] = loc
    # 暂存区
    first_shelf = next(iter(shelf_map.values()))
    staging = Location(shelf_id=first_shelf.id, code=STAGING_LOCATION,
                       shelf_no="STAGING", column_no="00", layer_no="00", status="empty")
    db.add(staging)
    loc_map[STAGING_LOCATION] = staging
    db.commit()
    for l in loc_map.values():
        db.refresh(l)
    print(f"库位：{len(loc_codes)} 个业务库位 + 1 个暂存区")

    # SKU（base + marker）
    sku_map = {}
    marker_count = 0
    for r in rows:
        code = str(r[0]).strip()
        stype = "marker" if code in MARKER_SKUS else "base"
        if stype == "marker":
            marker_count += 1
        name = r[4] if r[4] else None
        img = r[5] if len(r) > 5 and r[5] else None
        sku = Sku(sku_code=code, sku_name=name, sku_type=stype, image_url=img,
                  unit="pcs", status=1)
        db.add(sku)
        sku_map[code] = sku
    db.commit()
    for s in sku_map.values():
        db.refresh(s)
    print(f"SKU：{len(sku_map)} 个（含 {marker_count} 个标记 SKU）")

    # 期初库存
    stock_count = 0
    total_qty = 0
    staging_count = 0
    for i, r in enumerate(rows):
        code = str(r[0]).strip()
        qty = int(r[3])
        if code in MARKER_SKUS or qty <= 0:
            continue
        loc_code = r[2] if r[2] and r[2] != "无货架位" else None
        if loc_code is None:
            loc = loc_map[STAGING_LOCATION]
            staging_count += 1
        else:
            loc = loc_map[loc_code]
        sku = sku_map[code]
        stock = Stock(sku_id=sku.id, location_id=loc.id, batch_no=None,
                      available_qty=qty, allocated_qty=0, locked_qty=0, in_transit_qty=0)
        db.add(stock)
        db.flush()
        db.add(StockTransaction(stock_id=stock.id, sku_id=sku.id, change_type="inbound",
                                change_qty=qty, before_qty=0, after_qty=qty,
                                ref_type="INIT", ref_id=i + 1, remark="期初导入"))
        stock_count += 1
        total_qty += qty
    db.commit()
    print(f"期初库存：{stock_count} 条 stock 行，真实库存合计 {total_qty} 件（其中暂存区 {staging_count} 条）")

    db.execute(text("UPDATE location SET status='occupied' WHERE id IN (SELECT DISTINCT location_id FROM stock)"))
    db.commit()
    print("库位状态已更新（有库存 -> occupied）")


def import_file_b(db):
    cols, rows = read_rows(FILE_B, header_row=2)
    print(f"文件B：{len(rows)} 行，列={cols}")

    combo_map = {}
    for r in rows:
        code = str(r[0]).strip()
        name = r[4] if r[4] else None
        img = r[5] if len(r) > 5 and r[5] else None
        sku = Sku(sku_code=code, sku_name=name, sku_type="combo", image_url=img,
                  unit="set", status=1)
        db.add(sku)
        combo_map[code] = sku
    db.commit()
    for s in combo_map.values():
        db.refresh(s)
    print(f"组合 SKU：{len(combo_map)} 个")

    all_skus = {s.sku_code: s for s in db.query(Sku).all()}

    bom_count = 0
    miss = 0
    for r in rows:
        combo_code = str(r[0]).strip()
        bom_str = r[6] if len(r) > 6 and r[6] else ""
        if not bom_str:
            continue
        for seg in [s for s in bom_str.split(";") if s.strip()]:
            comp_code, qty = parse_bom_segment(seg)
            if comp_code not in all_skus:
                miss += 1
                continue
            db.add(Bom(combo_sku_id=all_skus[combo_code].id,
                       component_sku_id=all_skus[comp_code].id, qty=qty))
            bom_count += 1
    db.commit()
    print(f"BOM：{bom_count} 条（未匹配子SKU {miss} 个）")


def main():
    db = SessionLocal()
    try:
        print("=" * 60)
        print("T3.3 期初数据切换（对齐数据字典）")
        print("=" * 60)
        clear_init_data(db)
        fix_rules(db)
        import_file_a(db)
        import_file_b(db)
        print("=" * 60)
        print("期初切换完成")
        print("=" * 60)
    finally:
        db.close()


if __name__ == "__main__":
    main()
